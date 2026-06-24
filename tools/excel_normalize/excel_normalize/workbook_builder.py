"""入力用 Excel（構成 + ユニットシート）の生成。"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from .constants import FLOW_HEADERS, KOSEI_HEADERS, KOSEI_SHEET

INTERNAL_CODE = "Z00001"
DISPLAY_NAME = "プレス機B"

KOSEI_ROWS = [
    (INTERNAL_CODE, DISPLAY_NAME, "供給ユニット", "取出"),
    (INTERNAL_CODE, DISPLAY_NAME, "供給ユニット", "供給"),
    (INTERNAL_CODE, DISPLAY_NAME, "収納ユニット", "取出"),
    (INTERNAL_CODE, DISPLAY_NAME, "収納ユニット", "収納"),
]

FLOW_SAMPLES: dict[str, list[list[str]]] = {
    "取出": [
        ["10", "端子", "20", "", "1", "0", "開始", "", "", ""],
        ["20", "処理", "30", "", "2", "0", "ワーク取出", "", "", ""],
        ["30", "端子", "", "", "3", "0", "終了", "", "", ""],
    ],
    "供給": [
        ["10", "端子", "20", "", "1", "0", "開始", "", "", ""],
        ["20", "処理", "30", "", "2", "0", "供給実行", "", "", ""],
        ["30", "端子", "", "", "3", "0", "終了", "", "", ""],
    ],
    "収納": [
        ["10", "端子", "20", "", "1", "0", "開始", "", "", ""],
        ["20", "処理", "30", "", "2", "0", "収納実行", "", "", ""],
        ["30", "端子", "", "", "3", "0", "終了", "", "", ""],
    ],
}

USAGE_LINES = [
    "入力用 Excel テンプレ v0.1",
    "",
    "1. このファイルをコピーし、装置ごとに 1 ファイルを用意する（1 ファイル = 1 社内番号）。",
    "2. 「構成」シート: 装置製番・装置名・ユニット・動作を行順に記入（左ナビの並び）。",
    "3. ユニットごとにシートを追加し、シート名を構成のユニット列と完全一致させる。",
    "4. 各動作は「挿入 → テーブル」で 10 列の表にする。同一シート内で横並び可。",
    "5. テーブル名は「{ユニット短名}_{動作名}」（例: 供給_取出）。ブック全体で一意。",
    "6. 結合セルはテーブル内で使わない。列の意味は specs/03_技術仕様/列の意味.md を参照。",
    "",
    "正規化（import.json 生成）:",
    "  python -m excel_normalize.cli 入力用.xlsx -o import.json",
    "",
    "詳細: tools/excel_normalize/README.md",
    "      specs/03_技術仕様/Excel取込_正規化パイプライン.md",
]


def _add_flow_table(
    ws,
    *,
    table_name: str,
    start_col: int,
    start_row: int,
    data_rows: list[list[str]],
) -> None:
    headers = list(FLOW_HEADERS)
    row_count = 1 + len(data_rows)
    end_col = start_col + len(headers) - 1
    end_row = start_row + row_count - 1

    for c, header in enumerate(headers):
        ws.cell(start_row, start_col + c, header)
    for r_idx, data in enumerate(data_rows, start=1):
        for c_idx, value in enumerate(data):
            ws.cell(start_row + r_idx, start_col + c_idx, value)

    start_cell = ws.cell(start_row, start_col).coordinate
    end_cell = ws.cell(end_row, end_col).coordinate
    ref = f"{start_cell}:{end_cell}"

    safe_name = table_name.replace(" ", "_")
    table = Table(displayName=safe_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _populate_kosei_sheet(ws) -> None:
    ws.append(list(KOSEI_HEADERS))
    for row in KOSEI_ROWS:
        ws.append(list(row))


def _add_unit_sheets(wb: Workbook) -> None:
    ws_supply = wb.create_sheet("供給ユニット")
    _add_flow_table(
        ws_supply,
        table_name="供給_取出",
        start_col=1,
        start_row=1,
        data_rows=FLOW_SAMPLES["取出"],
    )
    _add_flow_table(
        ws_supply,
        table_name="供給_供給",
        start_col=13,
        start_row=1,
        data_rows=FLOW_SAMPLES["供給"],
    )

    ws_storage = wb.create_sheet("収納ユニット")
    _add_flow_table(
        ws_storage,
        table_name="収納_取出",
        start_col=1,
        start_row=1,
        data_rows=FLOW_SAMPLES["取出"],
    )
    _add_flow_table(
        ws_storage,
        table_name="収納_収納",
        start_col=13,
        start_row=1,
        data_rows=FLOW_SAMPLES["収納"],
    )


def _populate_usage_sheet(ws) -> None:
    """正規化対象外（先頭 `_`）の作者向け説明シート。"""
    ws.column_dimensions["A"].width = 88
    bold = Font(bold=True)
    for row_idx, line in enumerate(USAGE_LINES, start=1):
        cell = ws.cell(row_idx, 1, line)
        if row_idx == 1:
            cell.font = bold


def build_workbook(*, include_usage_sheet: bool = False) -> Workbook:
    wb = Workbook()
    default_ws = wb.active

    if include_usage_sheet:
        default_ws.title = "_使い方"
        _populate_usage_sheet(default_ws)
        ws_kosei = wb.create_sheet(KOSEI_SHEET)
    else:
        ws_kosei = default_ws
        ws_kosei.title = KOSEI_SHEET

    _populate_kosei_sheet(ws_kosei)
    _add_unit_sheets(wb)
    return wb
