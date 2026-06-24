"""装置マスター xlsx の共通生成（テンプレ · DEMO seed · 新規装置スキャフォールド）。"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font

from .constants import KOSEI_HEADERS, KOSEI_SHEET
from .workbook_builder import _add_flow_table

TABLE_GAP_COLS = 13


@dataclass(frozen=True)
class ModuleSpec:
    label: str
    process_text: str | None = None


@dataclass(frozen=True)
class UnitSpec:
    label: str
    modules: tuple[ModuleSpec, ...]


@dataclass(frozen=True)
class DeviceSpec:
    internal_code: str
    display_name: str
    units: tuple[UnitSpec, ...]


def unit_short(label: str) -> str:
    return label.removesuffix("ユニット")


def flow_rows(module: ModuleSpec) -> list[list[str]]:
    middle = module.process_text or f"{module.label}実行"
    return [
        ["10", "端子", "20", "", "1", "0", "開始", "", "", ""],
        ["20", "処理", "30", "", "2", "0", middle, "", "", ""],
        ["30", "端子", "", "", "3", "0", "終了", "", "", ""],
    ]


# DEMO-003（供給 + 加工 · 各 2 動作）と同型 — 現在の標準構成
DEFAULT_DEVICE_TEMPLATE = DeviceSpec(
    "NEW-001",
    "（装置名）",
    (
        UnitSpec(
            "供給ユニット",
            (ModuleSpec("取出", "ワーク取出"), ModuleSpec("供給")),
        ),
        UnitSpec(
            "加工ユニット",
            (ModuleSpec("プレス"), ModuleSpec("離脱")),
        ),
    ),
)

USAGE_LINES_V02 = [
    "入力用 Excel テンプレ v0.2（供給 + 加工 · 各 2 動作）",
    "",
    "1. このファイルをコピーし、装置フォルダに {社内番号}_{装置名}.xlsx として置く。",
    "   例: testdata/devices/NEW-001_（装置名）/NEW-001_（装置名）.xlsx",
    "   または npm run excel:new-device -- NEW-001 装置名",
    "2. 「構成」シート: 装置製番 · 装置名 · ユニット · 動作（左ナビの並び）。",
    "3. ユニットシート名は構成のユニット列と完全一致（供給ユニット · 加工ユニット）。",
    "4. 各動作は Excel テーブル（10 列）。同一シート内で横並び可。",
    "5. テーブル名: {ユニット短名}_{動作名}（例: 供給_取出 · 加工_プレス）。",
    "6. 結合セルはテーブル内で使わない。",
    "",
    "正規化:",
    "  python -m excel_normalize.cli NEW-001_（装置名）.xlsx -o import.json",
    "",
    "Web 取込: その他 → import.json を取込…",
    "詳細: tools/excel_normalize/README.md",
]


def _populate_usage_sheet(ws, lines: list[str]) -> None:
    ws.column_dimensions["A"].width = 88
    bold = Font(bold=True)
    for row_idx, line in enumerate(lines, start=1):
        cell = ws.cell(row_idx, 1, line)
        if row_idx == 1:
            cell.font = bold


def build_device_workbook(
    spec: DeviceSpec,
    *,
    include_usage_sheet: bool = False,
    usage_lines: list[str] | None = None,
) -> Workbook:
    wb = Workbook()
    default_ws = wb.active

    if include_usage_sheet:
        default_ws.title = "_使い方"
        _populate_usage_sheet(default_ws, usage_lines or USAGE_LINES_V02)
        ws_kosei = wb.create_sheet(KOSEI_SHEET)
    else:
        ws_kosei = default_ws
        ws_kosei.title = KOSEI_SHEET

    ws_kosei.append(list(KOSEI_HEADERS))
    for unit in spec.units:
        for module in unit.modules:
            ws_kosei.append(
                (spec.internal_code, spec.display_name, unit.label, module.label)
            )

    for unit in spec.units:
        ws_unit = wb.create_sheet(unit.label)
        short = unit_short(unit.label)
        start_col = 1
        for module in unit.modules:
            _add_flow_table(
                ws_unit,
                table_name=f"{short}_{module.label}",
                start_col=start_col,
                start_row=1,
                data_rows=flow_rows(module),
            )
            start_col += TABLE_GAP_COLS

    return wb
