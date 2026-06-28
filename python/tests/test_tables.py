from __future__ import annotations

from openpyxl import Workbook

from excel_normalize.kosei import KoseiRow, KoseiSheet
from excel_normalize.tables import extract_unit_sheet_tables
from excel_normalize.workbook_builder import FlowTableMeta, _add_flow_table


def _v03_kosei() -> KoseiSheet:
    rows = [
        KoseiRow(
            internal_code="A0001",
            display_name="塗布装置",
            unit_label="塗布1部",
            module_label="M016塗布1搬送1_取1(供給搬送2)",
            sort_index=0,
            uin_id=1,
            module_id=16,
        ),
        KoseiRow(
            internal_code="A0001",
            display_name="塗布装置",
            unit_label="塗布1部",
            module_label="M018塗布1搬送2_取3(ﾚﾍﾞﾘﾝｸﾞ)",
            sort_index=1,
            uin_id=1,
            module_id=18,
        ),
    ]
    return KoseiSheet(
        rows=rows,
        display_name="塗布装置",
        internal_code="A0001",
        format_version="v03",
    )


def test_meta_mid_resolves_despite_misleading_table_name() -> None:
    """コピペで 動作00018 等になっても、直上 MID で M016 に紐付く。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "U1"
    rows = [
        [
            "1",
            "端子",
            "",
            "2",
            "",
            "1",
            "0",
            "MR230000",
            "塗布1搬送1_取1(供給搬送2)_開始",
            "",
        ],
        ["2", "処理", "", "3", "", "2", "1", "MR230100", "RC013", "P03"],
    ]
    _add_flow_table(
        ws,
        table_name="動作00018",
        start_col=1,
        start_row=1,
        meta=FlowTableMeta(
            uin_id=1,
            unit_label="塗布1部",
            mid=16,
            module_label="M016塗布1搬送1_取1(供給搬送2)",
        ),
        data_rows=rows,
    )

    kosei = _v03_kosei()
    blocks = extract_unit_sheet_tables(
        ws, "塗布1部", kosei.modules_for_unit("塗布1部"), kosei
    )

    assert len(blocks) == 1
    assert blocks[0].module_label == "M016塗布1搬送1_取1(供給搬送2)"
    assert blocks[0].rows[0][8] == "塗布1搬送1_取1(供給搬送2)_開始"


def test_meta_with_blank_row_gap() -> None:
    """見出しと表の間に 1 行空き（作者手書きレイアウト）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "U1"
    ws.cell(1, 1, 1)
    ws.cell(1, 2, "塗布1部")
    ws.cell(2, 1, 16)
    ws.cell(2, 2, "M016塗布1搬送1_取1(供給搬送2)")
    _add_flow_table(
        ws,
        table_name="動作00018",
        start_col=1,
        start_row=4,
        data_rows=[
            ["1", "端子", "", "2", "", "1", "0", "開始", "", ""],
        ],
    )

    kosei = _v03_kosei()
    blocks = extract_unit_sheet_tables(
        ws, "塗布1部", kosei.modules_for_unit("塗布1部"), kosei
    )
    assert blocks[0].module_label.startswith("M016")
