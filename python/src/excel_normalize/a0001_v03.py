"""A0001 塗布装置 — v0.3 形式（10 ユニット × 10 モジュール）の Excel 生成。"""

from __future__ import annotations

from openpyxl import Workbook

from excel_normalize.constants import (
    KOSEI_HEADERS_V03,
    KOSEI_SHEET,
    TABLE_GAP_COLS,
)
from excel_normalize.workbook_builder import FlowTableMeta, _add_flow_table

INTERNAL_CODE = "A0001"
DISPLAY_NAME = "塗布装置"
UNIT_COUNT = 10
MODULES_PER_UNIT = 10
ROW_COUNT = UNIT_COUNT * MODULES_PER_UNIT


def unit_label(uin_id: int) -> str:
    return f"ﾕﾆｯﾄ{uin_id}"


def module_label(uin_id: int, index: int) -> str:
    return f"動作{uin_id * 100 + index:03d}"


def module_id(uin_id: int, index: int) -> int:
    return uin_id * 100 + index


# U0 手書き済み（動作008〜009 は import.json プレースホルダ · U1〜は未着手）
FLOW_TEXT_OVERRIDES: dict[str, str] = {
    "動作000": "ワーク取出",
}

# 動作001〜007 ≡ M001〜M007（作者 Excel 正本 · v2 列順）
FLOW_ROWS_OVERRIDES: dict[str, list[list[str]]] = {
    "動作001": [
        [
            "1",
            "端子",
            "",
            "2",
            "",
            "1",
            "0",
            "MR41500",
            "供給ﾏｶﾞｼﾞﾝ_SUS板取出位置_開始",
            "",
        ],
        ["2", "処理", "", "3", "", "2", "0", "MR41600", "RC008_供給ﾏｶﾞｼﾞﾝz", "P01_下降"],
        [
            "3",
            "処理",
            "",
            "4",
            "",
            "3",
            "0",
            "MR41602",
            "RC001_供給ﾏｶﾞｼﾞﾝy",
            "P02_SUS取出",
        ],
        ["4", "端子", "", "", "", "", "", "MR41501", "供給ﾏｶﾞｼﾞﾝ_SUS板取出位置_終了", ""],
    ],
    "動作002": [
        ["1", "端子", "", "2", "", "1", "0", "MR42000", "供給_SUS板搬送_取_開始", ""],
        [
            "2",
            "処理",
            "",
            "3,4,5",
            "",
            "2",
            "0",
            "MR42100",
            "RC008_供給ﾏｶﾞｼﾞﾝz",
            "P03_SUS取出下",
        ],
        ["3", "処理", "", "6", "", "3", "1", "MR42102", "供給SUSｽﾄｯﾊﾟｰ", "_出"],
        ["4", "処理", "", "6", "", "4", "2", "MR42104", "供給SUS板引供給ｼｬｯﾀｰ", "_開"],
        [
            "5",
            "処理",
            "",
            "6",
            "",
            "5",
            "3",
            "MR42106",
            "RC002_供給SUS板引y",
            "P02_SUS取出",
        ],
        ["6", "処理", "", "7", "", "6", "0", "MR42108", "SUS板取", "準備OK"],
        [
            "7",
            "処理",
            "",
            "8",
            "",
            "7",
            "0",
            "MR42200",
            "RC003_供給SUS板引x",
            "P02_SUS取出",
        ],
        [
            "8",
            "処理",
            "",
            "9",
            "",
            "8",
            "0",
            "MR42202",
            "RC008_供給ﾏｶﾞｼﾞﾝz",
            "P02_SUS取出上",
        ],
        [
            "9",
            "処理",
            "",
            "10",
            "",
            "9",
            "0",
            "MR42204",
            "供給SUS板引出ﾁｬｯｸ左右",
            "_閉",
        ],
        [
            "10",
            "処理",
            "",
            "11",
            "",
            "10",
            "0",
            "MR42206",
            "SUS板取完了",
            "_フラグシフト",
        ],
        [
            "11",
            "処理",
            "",
            "12,13",
            "",
            "11",
            "0",
            "MR42300",
            "RC003_供給SUS板引x",
            "P01_戻",
        ],
        ["12", "処理", "", "14", "", "12", "1", "MR42302", "供給SUSｽﾄｯﾊﾟｰ", "_戻"],
        [
            "13",
            "処理",
            "",
            "14",
            "",
            "13",
            "2",
            "MR42304",
            "供給SUS板引供給ｼｬｯﾀｰ",
            "_閉",
        ],
        ["14", "端子", "", "", "", "14", "0", "MR42001", "供給_SUS板搬送_取_終了", ""],
    ],
    "動作003": [
        ["1", "端子", "", "2", "", "1", "0", "MR42500", "供給_SUS板搬送_置_開始", ""],
        [
            "2",
            "処理",
            "",
            "3,4,5",
            "",
            "2",
            "0",
            "MR42600",
            "RC008_供給ﾏｶﾞｼﾞﾝz",
            "P02_SUS取出上",
        ],
        ["3", "処理", "", "6", "", "3", "1", "MR42602", "供給SUSｽﾄｯﾊﾟｰ", "_出"],
        ["4", "処理", "", "6", "", "4", "2", "MR42604", "供給SUS板引供給ｼｬｯﾀｰ", "_開"],
        [
            "5",
            "処理",
            "",
            "6",
            "",
            "5",
            "3",
            "MR42606",
            "RC002_供給SUS板引y",
            "P02_SUS取出",
        ],
        ["6", "処理", "", "7", "", "6", "0", "MR42608", "SUS板置", "準備OK"],
        [
            "7",
            "処理",
            "",
            "8",
            "",
            "7",
            "0",
            "MR42700",
            "RC003_供給SUS板引x",
            "P02_SUS取出",
        ],
        [
            "8",
            "処理",
            "",
            "9",
            "",
            "8",
            "0",
            "MR42702",
            "RC008_供給ﾏｶﾞｼﾞﾝz",
            "P03_SUS取出下",
        ],
        [
            "9",
            "処理",
            "",
            "10",
            "",
            "9",
            "0",
            "MR42704",
            "供給SUS板引出ﾁｬｯｸ左右",
            "_開",
        ],
        [
            "10",
            "処理",
            "",
            "11",
            "",
            "10",
            "0",
            "MR42706",
            "SUS板取完了",
            "_フラグシフト",
        ],
        [
            "11",
            "処理",
            "",
            "12,13",
            "",
            "11",
            "0",
            "MR42800",
            "RC003_供給SUS板引x",
            "P01_戻",
        ],
        ["12", "処理", "", "14", "", "12", "1", "MR42802", "供給SUSｽﾄｯﾊﾟｰ", "_戻"],
        [
            "13",
            "処理",
            "",
            "14",
            "",
            "13",
            "2",
            "MR42804",
            "供給SUS板引供給ｼｬｯﾀｰ",
            "_閉",
        ],
        ["14", "端子", "", "", "", "14", "0", "MR42501", "供給_SUS板搬送_置_終了", ""],
    ],
    "動作004": [
        ["1", "端子", "", "2", "", "1", "0", "MR43000", "供給_SUS板搬送_検査_開始", ""],
        [
            "2",
            "処理",
            "",
            "3",
            "",
            "2",
            "0",
            "MR43100",
            "RC002_供給SUS板引y",
            "P03_検査位置",
        ],
        ["3", "処理", "", "4", "", "3", "0", "MR43200", "SUS検査", "計測"],
        ["4", "判断", "", "6", "5", "", "0", "MR43204", "SUS検査", "結果確認"],
        ["5", "処理", "", "6", "", "", "1", "MR43208", "ALXXX", "画像検査異常"],
        ["6", "処理", "", "7", "", "", "0", "MR43210", "SUS検査", "_フラグシフト"],
        ["7", "端子", "", "", "", "", "0", "MR43001", "供給_SUS板搬送_検査_終了", ""],
    ],
    "動作005": [
        [
            "1",
            "端子",
            "",
            "2,3",
            "",
            "0",
            "0",
            "MR43500",
            "供給_個片搬送1(取出)_取_開始",
            "",
        ],
        [
            "2",
            "処理",
            "",
            "4",
            "",
            "1",
            "1",
            "MR43600",
            "RC002_供給SUS板引y",
            "P05_受渡",
        ],
        [
            "3",
            "処理",
            "",
            "4",
            "",
            "2",
            "2",
            "MR43602",
            "RC003_供給SUS板引x",
            "P05_受渡",
        ],
        ["4", "処理", "", "5", "", "0", "0", "MR43604", "個片取", "準備OK"],
        [
            "5",
            "処理",
            "",
            "6",
            "",
            "0",
            "0",
            "MR43701",
            "RC004_供給個片移載z",
            "P02_下降(SUS板)",
        ],
        [
            "6",
            "処理",
            "",
            "7",
            "",
            "0",
            "0",
            "MR43703",
            "RC005_供給個片移載c",
            "P02_閉-1mm",
        ],
        [
            "7",
            "処理",
            "",
            "8",
            "",
            "0",
            "0",
            "MR43705",
            "RC005_供給個片移載c",
            "P03_閉",
        ],
        ["8", "処理", "", "9", "", "0", "0", "MR43707", "個片取出", "_フラグシフト"],
        [
            "9",
            "端子",
            "",
            "10",
            "",
            "0",
            "0",
            "MR43709",
            "RC004_供給個片移載z",
            "P01_上昇",
        ],
        [
            "10",
            "端子",
            "",
            "",
            "",
            "0",
            "0",
            "MR43501",
            "供給_個片搬送1(取出)_取_終了",
            "",
        ],
    ],
    "動作006": [
        [
            "1",
            "端子",
            "",
            "2,3,4",
            "",
            "1",
            "0",
            "MR44000",
            "供給搬送1_置_x001F__開始",
            "",
        ],
        [
            "2",
            "処理",
            "",
            "5",
            "",
            "2",
            "1",
            "MR44100",
            "RC002_供給SUS板引y",
            "P04_退避",
        ],
        [
            "3",
            "処理",
            "",
            "5",
            "",
            "3",
            "2",
            "MR44102",
            "RC006_供給個片y",
            "P01_受取(手前)",
        ],
        ["4", "処理", "", "5", "", "", "3", "MR44104", "RC007_供給個片r", "P01_受取"],
        ["5", "処理", "", "6", "", "", "0", "MR44106", "個片置", "準備OK"],
        [
            "6",
            "処理",
            "",
            "7",
            "",
            "",
            "0",
            "MR44200",
            "RC004_供給個片移載z",
            "P03_下降(供給個片)",
        ],
        ["7", "処理", "", "8", "", "", "0", "MR44202", "供給搬送2真空", "_吸着"],
        ["8", "処理", "", "9", "", "", "0", "MR44204", "RC005_供給個片移載c", "P01_開"],
        ["9", "処理", "", "10", "", "", "0", "MR44206", "個片取出", "_ﾌﾗｸﾞｼﾌﾄ"],
        [
            "10",
            "判断",
            "",
            "12",
            "11",
            "",
            "0",
            "MR44300",
            "供給搬送2真空",
            "_吸着確認",
        ],
        ["11", "処理", "", "12", "", "", "1", "MR44303", "ALXXX", "吸着ミス"],
        [
            "12",
            "処理",
            "",
            "",
            "",
            "",
            "",
            "MR44304",
            "RC004_供給個片移載z",
            "P01_上昇",
        ],
        ["13", "端子", "", "", "", "", "", "MR44001", "供給搬送1_置_x001F__終了", ""],
    ],
    # 動作007 ≡ M007 — M006 置完了後の搬送1原点復帰（真空解放 → 個片軸退避 → SUS引戻）
    "動作007": [
        ["1", "端子", "", "2", "", "1", "0", "MR44500", "供給_個片搬送1(戻)_開始", ""],
        ["2", "処理", "", "3", "", "2", "0", "MR44600", "供給搬送2真空", "_解放"],
        ["3", "処理", "", "4", "", "3", "0", "MR44602", "RC007_供給個片r", "P02_退避"],
        ["4", "処理", "", "5", "", "4", "0", "MR44604", "RC006_供給個片y", "P02_退避"],
        ["5", "処理", "", "6", "", "5", "0", "MR44700", "RC003_供給SUS板引x", "P01_戻"],
        [
            "6",
            "処理",
            "",
            "7",
            "",
            "6",
            "0",
            "MR44702",
            "RC002_供給SUS板引y",
            "P04_待機",
        ],
        ["7", "処理", "", "8", "", "7", "0", "MR44704", "個片搬送1戻", "_フラグシフト"],
        ["8", "端子", "", "", "", "8", "0", "MR44501", "供給_個片搬送1(戻)_終了", ""],
    ],
}


def flow_rows_for_module(module_name: str) -> list[list[str]]:
    if module_name in FLOW_ROWS_OVERRIDES:
        return FLOW_ROWS_OVERRIDES[module_name]
    middle = FLOW_TEXT_OVERRIDES.get(module_name, module_name)
    return [
        ["10", "端子", "", "20", "", "1", "0", "開始", "", ""],
        ["20", "処理", "", "30", "", "2", "0", middle, "", ""],
        ["30", "端子", "", "", "", "3", "0", "終了", "", ""],
    ]


def kosei_row_formulas(excel_row: int) -> tuple[str, str, str, str, str, str]:
    """構成シート1行分の数式（マスター表を XLOOKUP）。"""
    return (
        '=XLOOKUP("*",装置名!$A$2:$A$2,装置名!$A$2:$A$2,,2)',
        '=XLOOKUP("*",装置名!$A$2:$A$2,装置名!$B$2:$B$2,,2)',
        "=QUOTIENT(ROW()-2,10)",
        f"=XLOOKUP(C{excel_row},ユニット!$A:$A,ユニット!$B:$B)",
        f"=C{excel_row}*100+MOD(ROW()-2,10)",
        f"=XLOOKUP(E{excel_row},モジュール!$A:$A,モジュール!$B:$B)",
    )


def _populate_master_sheets(wb: Workbook) -> None:
    ws_device = wb.create_sheet("装置名")
    ws_device.append(["装置製番", "装置名"])
    ws_device.append([INTERNAL_CODE, DISPLAY_NAME])

    ws_units = wb.create_sheet("ユニット")
    ws_units.append(["ID", "ユニット名", "MID数", "MID開始", "MID終了"])
    for uin_id in range(UNIT_COUNT):
        mid_start = uin_id * 100
        mid_end = mid_start + MODULES_PER_UNIT - 1
        ws_units.append(
            [uin_id, unit_label(uin_id), MODULES_PER_UNIT, mid_start, mid_end]
        )

    ws_modules = wb.create_sheet("モジュール")
    ws_modules.append(["ID", "動作"])
    for uin_id in range(UNIT_COUNT):
        for mod_index in range(MODULES_PER_UNIT):
            ws_modules.append(
                [module_id(uin_id, mod_index), module_label(uin_id, mod_index)]
            )


def _populate_kosei_sheet(wb: Workbook) -> None:
    ws_kosei = wb.create_sheet(KOSEI_SHEET, 0)
    ws_kosei.append(list(KOSEI_HEADERS_V03))
    for idx in range(ROW_COUNT):
        excel_row = idx + 2
        for col, formula in enumerate(kosei_row_formulas(excel_row), start=1):
            ws_kosei.cell(excel_row, col, formula)


def _populate_unit_flow_sheets(wb: Workbook) -> None:
    for uin_id in range(UNIT_COUNT):
        ws_unit = wb.create_sheet(f"U{uin_id}")
        start_col = 1
        for mod_index in range(MODULES_PER_UNIT):
            mod_name = module_label(uin_id, mod_index)
            mid = module_id(uin_id, mod_index)
            _add_flow_table(
                ws_unit,
                table_name=f"M{mid:03d}",
                start_col=start_col,
                start_row=1,
                meta=FlowTableMeta(
                    uin_id=uin_id,
                    unit_label=unit_label(uin_id),
                    mid=mid,
                    module_label=mod_name,
                ),
                data_rows=flow_rows_for_module(mod_name),
            )
            start_col += TABLE_GAP_COLS


def build_a0001_v03_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _populate_master_sheets(wb)
    _populate_kosei_sheet(wb)
    _populate_unit_flow_sheets(wb)
    return wb
