from __future__ import annotations

import re
from dataclasses import dataclass

from openpyxl.utils import range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .constants import FLOW_COLUMN_COUNT, FLOW_HEADERS, RESERVED_SHEET_NAMES
from .kosei import KoseiSheet

# A0001 等の短縮列ヘッダー → 正規化コード列名（§6.3）
FLOW_HEADER_ALIASES: dict[str, str] = {
    "種別": "図形種別",
    "下先": "接続先(下)",
    "文1": "Text1",
    "文2": "Text2",
    "文3": "Text3",
}


@dataclass(frozen=True)
class FlowTableBlock:
    table_name: str
    module_label: str
    rows: list[list[str]]


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _unit_short_label(unit_label: str) -> str:
    if unit_label.endswith("ユニット"):
        return unit_label[: -len("ユニット")]
    return unit_label


def _normalize_header(h: str) -> str:
    """短縮ヘッダーを正規化コード列名に変換。"""
    return FLOW_HEADER_ALIASES.get(h, h)


def _mid_from_label(label: str) -> int | None:
    """'M016 供給...' · 'M000供給...' 等のモジュールラベルから MID を抽出。"""
    m = re.match(r"^M(\d+)", label)
    return int(m.group(1)) if m else None


def _parse_int_cell(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if text.isdigit():
        return int(text)
    return None


def _read_table_meta_rows(
    ws: Worksheet, header_row: int, min_col: int
) -> tuple[int, int] | None:
    """テーブルヘッダー行直上の UinID / MID 行を読む（1 行空きにも対応）。"""
    for uin_off, mid_off in ((-2, -1), (-3, -2)):
        uin_row = header_row + uin_off
        mid_row = header_row + mid_off
        if uin_row < 1 or mid_row < 1:
            continue
        uin_id = _parse_int_cell(ws.cell(uin_row, min_col).value)
        mid = _parse_int_cell(ws.cell(mid_row, min_col).value)
        if uin_id is not None and mid is not None:
            return uin_id, mid
    return None


def _resolve_module_label(
    ws: Worksheet,
    table_name: str,
    header_row: int,
    min_col: int,
    unit_label: str,
    expected_modules: list[str],
    kosei: KoseiSheet,
) -> str | None:
    meta = _read_table_meta_rows(ws, header_row, min_col)
    if meta is not None:
        uin_id, mid = meta
        label = kosei.module_label_for_mid(unit_label, mid, uin_id=uin_id)
        if label is not None and label in expected_modules:
            return label
        return None

    return resolve_table_module_label(table_name, unit_label, expected_modules)


def resolve_table_module_label(
    table_name: str,
    unit_label: str,
    expected_modules: list[str],
) -> str | None:
    # 完全一致
    if table_name in expected_modules:
        return table_name

    short = _unit_short_label(unit_label)
    prefixed = f"{short}_{table_name}"
    if prefixed in expected_modules:
        return table_name

    for mod in expected_modules:
        if table_name == f"{short}_{mod}":
            return mod
        if table_name == f"{unit_label}_{mod}":
            return mod

    # v0.3: テーブル名 動作N の数値部分を MID として解釈しラベルプレフィックスと照合
    # 例: 動作000 → 0 · 動作00018 → 18 · 動作00119 → 19（%100 フォールバック）
    mid_match = re.match(r"^動作(\d+)$", table_name)
    if mid_match:
        n = int(mid_match.group(1))
        mid_candidates = [n]
        if n >= 100:
            mid_candidates.append(n % 100)
        for target_mid in mid_candidates:
            for mod in expected_modules:
                if _mid_from_label(mod) == target_mid:
                    return mod

    return None


def _is_header_row(row: list[str]) -> bool:
    if len(row) < FLOW_COLUMN_COUNT:
        return False
    head = [_normalize_header(c.strip()) for c in row[:FLOW_COLUMN_COUNT]]
    if head == list(FLOW_HEADERS):
        return True
    first = head[0].lower()
    return first in ("id", "ｉｄ")


def _read_table_rows(ws: Worksheet, ref: str) -> list[list[str]]:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    matrix: list[list[str]] = []
    for r in range(min_row, max_row + 1):
        row: list[str] = []
        for c in range(min_col, min_col + FLOW_COLUMN_COUNT):
            row.append(_cell_str(ws.cell(r, c).value))
        matrix.append(row)
    return matrix


def _data_rows_from_matrix(matrix: list[list[str]]) -> list[list[str]]:
    if not matrix:
        return []
    start = 1 if _is_header_row(matrix[0]) else 0
    data: list[list[str]] = []
    for row in matrix[start:]:
        if not any(cell for cell in row):
            continue
        if len(row) < FLOW_COLUMN_COUNT:
            raise ValueError(
                f"10列必要ですが {len(row)} 列です（先頭セル: {row[0] if row else '空'}）"
            )
        normalized = row[:FLOW_COLUMN_COUNT]
        while len(normalized) < FLOW_COLUMN_COUNT:
            normalized.append("")
        data.append(normalized)
    return data


def extract_unit_sheet_tables(
    ws: Worksheet,
    unit_label: str,
    expected_modules: list[str],
    kosei: KoseiSheet,
) -> list[FlowTableBlock]:
    if not ws.tables:
        raise ValueError(
            f"シート「{ws.title}」: Excel テーブルがありません。"
            f"各動作を「挿入 → テーブル」で登録してください（期待動作: {', '.join(expected_modules)}）"
        )

    blocks: list[FlowTableBlock] = []
    seen_modules: set[str] = set()

    for table in ws.tables.values():
        table_name = table.name
        min_col, header_row, _, _ = range_boundaries(table.ref)
        matrix = _read_table_rows(ws, table.ref)
        try:
            data_rows = _data_rows_from_matrix(matrix)
        except ValueError as exc:
            raise ValueError(
                f"シート「{ws.title}」· テーブル「{table_name}」: {exc}"
            ) from exc

        if not data_rows:
            raise ValueError(
                f"シート「{ws.title}」· テーブル「{table_name}」: データ行がありません"
            )

        module_label = _resolve_module_label(
            ws,
            table_name,
            header_row,
            min_col,
            unit_label,
            expected_modules,
            kosei,
        )
        if module_label is None:
            meta = _read_table_meta_rows(ws, header_row, min_col)
            if meta is not None:
                uin_id, mid = meta
                raise ValueError(
                    f"シート「{ws.title}」· テーブル「{table_name}」: "
                    f"直上の MID {mid}（UinID {uin_id}）が"
                    f"構成のユニット「{unit_label}」にありません"
                )
            raise ValueError(
                f"シート「{ws.title}」· テーブル「{table_name}」: "
                f"構成シートの動作名と一致しません（期待: {', '.join(expected_modules)}）"
            )
        if module_label in seen_modules:
            raise ValueError(
                f"シート「{ws.title}」: 動作「{module_label}」に対応するテーブルが重複しています"
            )
        seen_modules.add(module_label)
        blocks.append(
            FlowTableBlock(
                table_name=table_name,
                module_label=module_label,
                rows=data_rows,
            )
        )

    return blocks


def list_unit_sheets(workbook: Workbook, kosei: KoseiSheet) -> dict[str, Worksheet]:
    sheets: dict[str, Worksheet] = {}
    for name in workbook.sheetnames:
        if name in RESERVED_SHEET_NAMES:
            continue
        if name.startswith("_"):
            continue
        sheets[name] = workbook[name]

    unit_sheets: dict[str, Worksheet] = {}
    for unit_label in kosei.unit_labels:
        sheet_title = kosei.unit_sheet_title(unit_label)
        if sheet_title not in sheets:
            continue
        unit_sheets[unit_label] = sheets[sheet_title]

    mapped_titles = {kosei.unit_sheet_title(u) for u in kosei.unit_labels}
    extra = set(sheets.keys()) - mapped_titles
    if extra:
        raise ValueError(
            f"構成に無いユニットシートがあります: {', '.join(sorted(extra))}"
        )

    return unit_sheets
