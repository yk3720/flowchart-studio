"""装置 xlsx の記入状況サマリー（正規化前の確認用）。"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from .constants import RESERVED_SHEET_NAMES
from .kosei import parse_kosei_sheet
from .tables import extract_unit_sheet_tables, list_unit_sheets


@dataclass(frozen=True)
class FlowEntry:
    unit_label: str
    module_label: str
    row_count: int


@dataclass
class InspectReport:
    workbook_path: Path
    internal_code: str
    display_name: str
    format_version: str
    sheet_names: list[str]
    unit_count: int
    module_count: int
    unit_sheets_expected: list[str]
    unit_sheets_present: list[str]
    unit_sheets_missing: list[str]
    flows_present: list[FlowEntry] = field(default_factory=list)
    flows_missing: list[tuple[str, str]] = field(default_factory=list)
    blockers: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def normalize_ready(self) -> bool:
        return not self.blockers

    def to_dict(self) -> dict:
        return {
            "workbook": str(self.workbook_path),
            "internal_code": self.internal_code,
            "display_name": self.display_name,
            "format_version": self.format_version,
            "sheet_names": self.sheet_names,
            "unit_count": self.unit_count,
            "module_count": self.module_count,
            "unit_sheets_expected": self.unit_sheets_expected,
            "unit_sheets_present": self.unit_sheets_present,
            "unit_sheets_missing": self.unit_sheets_missing,
            "flows_present": [
                {
                    "unit_label": f.unit_label,
                    "module_label": f.module_label,
                    "row_count": f.row_count,
                }
                for f in self.flows_present
            ],
            "flows_missing": [
                {"unit_label": u, "module_label": m} for u, m in self.flows_missing
            ],
            "flow_present_count": len(self.flows_present),
            "flow_missing_count": len(self.flows_missing),
            "normalize_ready": self.normalize_ready,
            "blockers": self.blockers,
            "warnings": self.warnings,
        }


def _unit_flow_sheets(workbook_sheet_names: list[str]) -> list[str]:
    return [
        name
        for name in workbook_sheet_names
        if name not in RESERVED_SHEET_NAMES and not name.startswith("_")
    ]


def inspect_workbook(workbook_path: Path) -> InspectReport:
    wb = load_workbook(workbook_path, data_only=True)
    kosei = parse_kosei_sheet(wb)

    expected_titles = [kosei.unit_sheet_title(u) for u in kosei.unit_labels]
    flow_sheets = _unit_flow_sheets(wb.sheetnames)
    present_titles = [t for t in expected_titles if t in wb.sheetnames]
    missing_titles = [t for t in expected_titles if t not in wb.sheetnames]

    report = InspectReport(
        workbook_path=workbook_path.resolve(),
        internal_code=kosei.internal_code,
        display_name=kosei.display_name,
        format_version=kosei.format_version,
        sheet_names=list(wb.sheetnames),
        unit_count=len(kosei.unit_labels),
        module_count=len(kosei.rows),
        unit_sheets_expected=expected_titles,
        unit_sheets_present=present_titles,
        unit_sheets_missing=missing_titles,
    )

    if missing_titles:
        report.warnings.append(
            f"ユニットシート未作成: {', '.join(missing_titles)}（段階手書きではスキップ）"
        )

    for row in kosei.rows:
        if "\x1f" in row.module_label or row.module_label.endswith("_x001F_"):
            report.warnings.append(
                f"モジュール名に制御文字 U+001F の痕跡: {row.unit_label} · {row.module_label}"
            )

    extra_flow_sheets = sorted(set(flow_sheets) - set(expected_titles))
    if extra_flow_sheets:
        report.blockers.append(
            f"構成に無いユニットシート: {', '.join(extra_flow_sheets)}"
        )

    try:
        unit_sheets = list_unit_sheets(wb, kosei)
    except ValueError as exc:
        report.blockers.append(str(exc))
        unit_sheets = {}

    found: set[tuple[str, str]] = set()
    for unit_label, ws in unit_sheets.items():
        expected_modules = kosei.modules_for_unit(unit_label)
        try:
            blocks = extract_unit_sheet_tables(ws, unit_label, expected_modules)
        except ValueError as exc:
            report.blockers.append(str(exc))
            continue
        for block in blocks:
            found.add((unit_label, block.module_label))
            report.flows_present.append(
                FlowEntry(
                    unit_label=unit_label,
                    module_label=block.module_label,
                    row_count=len(block.rows),
                )
            )

    for row in kosei.rows:
        key = (row.unit_label, row.module_label)
        if key not in found:
            report.flows_missing.append(key)

    if report.flows_missing:
        report.warnings.append(
            f"フロー未記入: {len(report.flows_missing)}/{report.module_count} モジュール"
        )

    return report


def format_inspect_report(report: InspectReport) -> str:
    lines = [
        f"{report.internal_code} {report.display_name}",
        f"  ファイル: {report.workbook_path.name}",
        f"  形式: {report.format_version} · シート: {', '.join(report.sheet_names)}",
        f"  構成: {report.unit_count} ユニット · {report.module_count} モジュール",
        f"  Uシート: あり {', '.join(report.unit_sheets_present) or '（なし）'}",
    ]
    if report.unit_sheets_missing:
        lines.append(f"           なし {', '.join(report.unit_sheets_missing)}")
    lines.append(f"  フロー記入: {len(report.flows_present)}/{report.module_count}")
    for entry in report.flows_present:
        lines.append(
            f"    · {entry.unit_label} · {entry.module_label}（{entry.row_count} 行）"
        )
    if report.blockers:
        lines.append("  normalize: ✗ ブロッカーあり")
        for msg in report.blockers:
            lines.append(f"    ! {msg}")
    else:
        lines.append("  normalize: ✓ 実行可（未記入フローは警告）")
    for msg in report.warnings:
        lines.append(f"  ⚠ {msg}")
    return "\n".join(lines)
