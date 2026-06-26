"""v0.3 マスターシート読み取り（構成の XLOOKUP 未評価時のフォールバック）。"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl import Workbook

from .labels import sanitize_module_label


@dataclass(frozen=True)
class UnitBand:
    """ユニットの MID 帯情報（§4.2）。"""

    uid: int
    label: str
    mid_count: int  # 0 = 未使用ユニット（プレースホルダ）またはバンド情報なし
    mid_start: int | None  # mid_count=0 のとき None
    mid_end: int | None  # mid_count=0 のとき None


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_int_cell(value: object) -> int | None:
    text = _cell_str(value)
    if not text:
        return None
    return int(float(text))


def read_v03_masters(
    workbook: Workbook,
) -> tuple[str, str, dict[int, UnitBand], dict[int, str]]:
    """装置名・ユニット（MID帯含む）・モジュールマスタを読む。

    Returns:
        (internal_code, display_name, units_map, modules_map)
        units_map: uid → UnitBand（MID数/開始/終了が揃っていればバンド情報あり）
        modules_map: MID → モジュール名
    """
    if "装置名" not in workbook.sheetnames:
        raise ValueError("シート「装置名」がありません（v0.3 マスター）")

    ws_device = workbook["装置名"]
    device_rows = list(ws_device.iter_rows(min_row=2, values_only=True))
    if not device_rows:
        raise ValueError("シート「装置名」にデータ行がありません")

    internal_code = _cell_str(device_rows[0][0])
    display_name = _cell_str(device_rows[0][1])
    if not internal_code or not display_name:
        raise ValueError("シート「装置名」2行目に装置製番・装置名を入力してください")

    units: dict[int, UnitBand] = {}
    if "ユニット" in workbook.sheetnames:
        for row in workbook["ユニット"].iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            uid = _parse_int_cell(row[0])
            label = _cell_str(row[1] if len(row) > 1 else None)
            if uid is None or not label:
                continue
            # v0.3: 列 2,3,4 = MID数, MID開始, MID終了（旧 2 列フォーマットでは 0/None）
            mid_count = _parse_int_cell(row[2] if len(row) > 2 else None) or 0
            mid_start: int | None = None
            mid_end: int | None = None
            if mid_count > 0:
                mid_start = _parse_int_cell(row[3] if len(row) > 3 else None)
                mid_end = _parse_int_cell(row[4] if len(row) > 4 else None)
            units[uid] = UnitBand(
                uid=uid,
                label=label,
                mid_count=mid_count,
                mid_start=mid_start,
                mid_end=mid_end,
            )

    modules: dict[int, str] = {}
    if "モジュール" in workbook.sheetnames:
        for row in workbook["モジュール"].iter_rows(min_row=2, values_only=True):
            mid = _parse_int_cell(row[0] if row else None)
            label = _cell_str(row[1] if row and len(row) > 1 else None)
            if mid is not None and label:
                modules[mid] = sanitize_module_label(label)

    return internal_code, display_name, units, modules
